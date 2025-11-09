"""
Router Matching - Lancement et résultats de matchings
"""

import json
import asyncio
import sys
from typing import List, Optional
from pathlib import Path
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from brainrh.paths import PROJECT_ROOT

# Ajouter le PROJECT_ROOT pour imports
sys.path.insert(0, str(PROJECT_ROOT))

from lib.models import Offre, MatchingResponse, ResultatMatching, MatchingMetadata
from matching_engine import MatchingEngine
from config_loader import Config

router = APIRouter()


class MatchingRequest(BaseModel):
    """Requête pour lancer un matching"""
    offre: Offre
    cvs: List[str]  # Liste d'IDs de CVs (noms de fichiers)
    top_n_rerank: int = 10
    model: str = "gpt-5-mini"
    concurrency: int = 500
    qps: float = 10.0


# Initialiser matching engine (singleton)
config = Config()
matching_engine = MatchingEngine(config._config)


@router.post("/run", response_model=MatchingResponse)
async def run_matching(request: MatchingRequest):
    """
    Lance un matching complet en mode batch

    **Pipeline:**
    1. Filtrage must-have (éliminatoire)
    2. Calcul similarité (embeddings)
    3. Détection nice-have manquants
    4. Re-ranking LLM du top N

    **Body:**
    - `offre`: Offre avec must-have/nice-have
    - `cvs`: Liste d'IDs de CVs (noms de fichiers parsés)
    - `top_n_rerank`: Nombre de CVs à re-ranker avec LLM (default: 10)
    - `model`: Modèle LLM (default: gpt-5-mini)
    - `concurrency`: Concurrence max (default: 500)
    - `qps`: QPS max (default: 10.0)

    **Retourne:** Résultats complets une fois terminé

    **Note:** Pour feedback temps-réel, utiliser `/matching/run/stream`
    """
    import time

    start_time = time.time()

    try:
        # Charger les CVs depuis cvs_parsed/
        # TODO: Gérer stockage persistant (pour l'instant on suppose que les CVs sont dans un dict)
        # Pour démonstration, on retourne une erreur si pas de système de stockage
        raise HTTPException(
            status_code=501,
            detail="Matching batch nécessite système de stockage persistant (CVs parsés). Utiliser /matching/run/stream avec données complètes."
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors du matching: {str(e)}"
        )


@router.get("/run/stream")
async def run_matching_stream(
    project_id: str = Query(..., description="ID du projet"),
    top_n_rerank: int = Query(default=10, ge=1, le=50, description="Nombre de CVs à re-ranker (max 50)"),
    model: str = Query(default="gpt-5-mini", description="Modèle LLM")
):
    """
    Lance un matching avec streaming SSE depuis un project_id

    **Query params:**
    - `project_id`: ID du projet (obligatoire)
    - `top_n_rerank`: Nombre de CVs à re-ranker (default: 10)
    - `model`: Modèle LLM (default: gpt-5-mini)

    **Événements:**
    - `progress`: Progression par étape (must_have_filtering, embedding, reranking)
    - `done`: Résultats finaux + métadonnées
    - `error`: Erreur

    **Pipeline complet:**
    1. Chargement de l'offre et des CVs depuis le projet
    2. Filtrage must-have (éliminatoire)
    3. Calcul similarité embeddings + nice-have
    4. Re-ranking LLM du top N
    5. Sauvegarde des résultats dans projects/{project_id}/matchings/{timestamp}/

    **Consommation côté client:**
    ```javascript
    const eventSource = new EventSource('/api/v1/matching/run/stream?project_id=X&top_n_rerank=10&model=gpt-5-mini');

    eventSource.addEventListener('progress', (e) => {
      const data = JSON.parse(e.data);
      console.log(`Étape: ${data.step}, ${data.current}/${data.total}`);
    });

    eventSource.addEventListener('done', (e) => {
      const data = JSON.parse(e.data);
      console.log('Matching terminé:', data.summary);
      eventSource.close();
    });
    ```
    """
    async def event_generator():
        import time
        import asyncio

        start_time = time.time()

        try:
            # === Charger le projet, l'offre et les CVs ===
            from project_manager import ProjectManager

            pm = ProjectManager(projects_folder="projects")

            # Charger le projet
            project = pm.get_project(project_id)
            if not project:
                yield "event: error\n"
                error_data = {'code': 'PROJECT_NOT_FOUND', 'message': f'Projet {project_id} introuvable'}
                yield f"data: {json.dumps(error_data)}\n\n"
                return

            # Charger l'offre
            offre_data = pm.load_offer(project_id)
            if not offre_data or not offre_data.get("sections"):
                yield "event: error\n"
                error_data = {'code': 'NO_OFFRE', 'message': "Aucune offre définie pour ce projet. Créez d'abord une offre."}
                yield f"data: {json.dumps(error_data)}\n\n"
                return

            # Charger les CVs parsés
            project_path = pm.get_project_path(project_id)
            if not project_path:
                yield "event: error\n"
                error_data = {'code': 'PROJECT_PATH_ERROR', 'message': f'Impossible de trouver le chemin du projet {project_id}'}
                yield f"data: {json.dumps(error_data)}\n\n"
                return

            cvs_parsed_dir = project_path / "cvs_parsed"
            if not cvs_parsed_dir.exists():
                yield "event: error\n"
                error_data = {'code': 'NO_CVS', 'message': "Aucun CV parsé pour ce projet. Parsez d'abord des CVs."}
                yield f"data: {json.dumps(error_data)}\n\n"
                return

            cv_files = list(cvs_parsed_dir.glob("*.json"))
            if not cv_files:
                yield "event: error\n"
                error_data = {'code': 'NO_CVS', 'message': 'Aucun CV parsé pour ce projet.'}
                yield f"data: {json.dumps(error_data)}\n\n"
                return

            # Charger les cv_meta pour obtenir file_path (CV originaux)
            from brainrh.services import CVService
            cv_metas = CVService.list_by_project(project_id)
            cv_meta_map = {meta["filename"]: meta.get("file_path") for meta in cv_metas}
            print(f"📇 {len(cv_meta_map)} entrées cv_meta chargées")

            # Charger les données CV
            cvs_data = []
            print(f"📂 Fichiers CV trouvés: {len(cv_files)}")
            for cv_file in cv_files:
                try:
                    with open(cv_file, 'r', encoding='utf-8') as f:
                        cv_data = json.load(f)
                        cvs_data.append(cv_data)
                        print(f"  ✅ Chargé: {cv_file.name}")
                except Exception as e:
                    print(f"  ❌ Erreur lecture CV {cv_file}: {e}")
                    continue

            total_cvs_initial = len(cvs_data)
            print(f"📊 Total CVs chargés en mémoire: {total_cvs_initial}")

            # Extraire must-have et nice-have de l'offre
            must_have = offre_data.get("must_have", [])
            nice_have = offre_data.get("nice_have", [])

            # Générer job_text pour embeddings (TOUTES les sections)
            sections = offre_data.get("sections", {})
            job_text = " ".join([
                str(v) for v in sections.values() if v
            ])

            # job_description = MÊME contenu que job_text pour alignement Streamlit
            job_description = job_text

            # === ÉTAPE 1: Filtrage must-have ===
            message = f'Filtrage must-have ({len(must_have)} critères)'
            yield "event: progress\n"
            yield f"data: {json.dumps({'step': 'must_have_filtering', 'current': 0, 'total': total_cvs_initial, 'progress': 0.0, 'message': message})}\n\n"

            # Callback pour progression
            def progress_callback_must_have(current, total):
                pass  # SSE progression déjà envoyée

            # Get event loop
            loop = asyncio.get_event_loop()

            # Exécuter filtrage must-have en thread pool (fonction sync)
            filtered_cvs = await loop.run_in_executor(
                None,
                lambda: matching_engine.filter_cvs_by_must_have(
                    cvs=cvs_data,
                    indispensables=must_have,
                    job_description=job_description,
                    use_parallel=True,
                    progress_callback=progress_callback_must_have
                )
            )

            total_cvs_filtered = len(filtered_cvs)

            yield "event: progress\n"
            progress_msg = f'Filtrage terminé: {total_cvs_filtered}/{total_cvs_initial} CVs acceptés'
            progress_data = {'step': 'must_have_filtering', 'current': total_cvs_initial, 'total': total_cvs_initial, 'progress': 1.0, 'message': progress_msg}
            yield f"data: {json.dumps(progress_data)}\n\n"

            if total_cvs_filtered == 0:
                yield "event: done\n"
                done_data = {'summary': {'results': [], 'metadata': {'total_cvs': total_cvs_initial, 'filtered_must_have': 0, 'top_reranked': 0, 'duree_totale_s': round(time.time() - start_time, 2)}}}
                yield f"data: {json.dumps(done_data)}\n\n"
                return

            # === ÉTAPE 2: Calcul similarité + Nice-have (compute_similarity_with_scoring fait les 2) ===
            yield "event: progress\n"
            embed_data = {'step': 'embedding', 'current': 0, 'total': total_cvs_filtered, 'progress': 0.0, 'message': 'Calcul des embeddings et nice-have'}
            yield f"data: {json.dumps(embed_data)}\n\n"

            def progress_callback_scoring(current, total):
                pass  # SSE déjà géré

            # Exécuter scoring en thread pool (fonction sync)
            scored_cvs = await loop.run_in_executor(
                None,
                lambda: matching_engine.compute_similarity_with_scoring(
                    job_text=job_text,
                    cvs=filtered_cvs,
                    nice_have_list=nice_have,
                    job_description=job_description,
                    progress_callback=progress_callback_scoring
                )
            )

            yield "event: progress\n"
            embed_done_data = {'step': 'embedding', 'current': total_cvs_filtered, 'total': total_cvs_filtered, 'progress': 1.0, 'message': 'Scoring et nice-have terminés'}
            yield f"data: {json.dumps(embed_done_data)}\n\n"

            # === ÉTAPE 3: Re-ranking LLM ===
            top_n = min(top_n_rerank, len(scored_cvs))
            yield "event: progress\n"
            rerank_msg = f'Re-ranking top {top_n}'
            rerank_data = {'step': 'reranking', 'current': 0, 'total': top_n, 'progress': 0.0, 'message': rerank_msg}
            yield f"data: {json.dumps(rerank_data)}\n\n"

            def progress_callback_rerank(current, total):
                pass

            # Exécuter re-ranking en thread pool avec le top_n de l'utilisateur
            reranked_cvs = await loop.run_in_executor(
                None,
                lambda: matching_engine.rerank_with_llm(
                    top_cvs=scored_cvs,
                    job_description=job_description,
                    progress_callback=progress_callback_rerank,
                    top_n=top_n  # Passer le paramètre utilisateur
                )
            )

            yield "event: progress\n"
            rerank_done_data = {'step': 'reranking', 'current': top_n, 'total': top_n, 'progress': 1.0, 'message': 'Re-ranking terminé'}
            yield f"data: {json.dumps(rerank_done_data)}\n\n"

            # === Événement done ===
            total_duration = time.time() - start_time

            # Convertir les résultats en format API
            # CORRECTION: Fusionner reranked_cvs avec scored_cvs pour préserver les scores
            results = []

            # Créer un dictionnaire des scores originaux indexé par nom de CV
            scored_map = {cv.get("cv"): cv for cv in scored_cvs}

            for cv_result in reranked_cvs:
                cv_name = cv_result.get("cv", "inconnu")

                # Récupérer les données de scoring originales
                original_data = scored_map.get(cv_name, {})

                # Récupérer le coefficient d'expérience
                coefficient_experience = cv_result.get("coefficient_qualite_experience", 1.0)

                # Recalculer le score_final avec le coefficient d'expérience
                score_base = original_data.get("score_base", 0.0)
                bonus_nice_have = original_data.get("bonus_nice_have_multiplicateur", 1.0)
                score_final_with_coef = score_base * bonus_nice_have * coefficient_experience
                score_final_with_coef = max(0.0, min(1.0, score_final_with_coef))

                # Construire les objets Evidence et Flags à partir des données brutes
                evidences_raw = cv_result.get("evidences") or []
                evidences = []
                for ev in evidences_raw:
                    if isinstance(ev, dict):
                        evidences.append({
                            "id": ev.get("id", ""),
                            "type": ev.get("type", "section"),
                            "ref": ev.get("ref", "")
                        })

                # Evidence map (normaliser None → {})
                evidence_map_raw = cv_result.get("evidence_map") or {}
                evidence_map = {
                    "commentaire_scoring": evidence_map_raw.get("commentaire_scoring", []),
                    "appreciation_globale": evidence_map_raw.get("appreciation_globale", [])
                }

                # Flags (normaliser None → {})
                flags_raw = cv_result.get("flags_raw") or {}
                flags = {
                    "gappes": flags_raw.get("gappes", []),
                    "overlaps": flags_raw.get("overlaps", [])
                }

                # URL vers le CV original (PDF/DOCX)
                # Mapping intelligent pour gérer les différences d'extensions
                original_cv_url = None
                file_path = None

                # Essai 1: correspondance directe
                if cv_name in cv_meta_map:
                    file_path = cv_meta_map[cv_name]

                # Essai 2: ajouter .json si manquant
                if not file_path and not cv_name.endswith('.json'):
                    json_name = cv_name + '.json'
                    if json_name in cv_meta_map:
                        file_path = cv_meta_map[json_name]

                # Essai 3: remplacer extension par .json
                if not file_path and '.' in cv_name:
                    from pathlib import Path
                    base_name = Path(cv_name).stem
                    json_name = base_name + '.json'
                    if json_name in cv_meta_map:
                        file_path = cv_meta_map[json_name]

                if file_path:
                    original_cv_url = f"/api/v1/files/{file_path}"

                results.append({
                    "cv": cv_name,
                    # Score final recalculé avec coefficient d'expérience
                    "score_final": score_final_with_coef,
                    "score_base": score_base,
                    "bonus_nice_have_multiplicateur": bonus_nice_have,
                    "nice_have_manquants": original_data.get("nice_have_manquants", []),
                    # Récupérer le coefficient et les commentaires depuis reranked_cvs
                    "coefficient_qualite_experience": coefficient_experience,
                    "commentaire_scoring": cv_result.get("commentaire_scoring", ""),
                    "appreciation_globale": cv_result.get("appreciation_globale", ""),
                    # Nouveaux champs
                    "evidences": evidences,
                    "evidence_map": evidence_map if evidence_map_raw else None,
                    "flags": flags if flags_raw else None,
                    "original_cv_url": original_cv_url
                })

            # Trier les résultats par score_final décroissant (avec coefficient appliqué)
            results.sort(key=lambda x: x["score_final"], reverse=True)

            summary = {
                "results": results,
                "metadata": {
                    "total_cvs": total_cvs_initial,
                    "filtered_must_have": total_cvs_filtered,
                    "top_reranked": len(reranked_cvs),
                    "duree_totale_s": round(total_duration, 2),
                    "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
                }
            }

            # Sauvegarder les résultats dans le projet (utilise le bon chemin: projects/ ou enterprises/)
            timestamp_str = time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime(start_time))
            matchings_dir = project_path / "matchings" / timestamp_str
            matchings_dir.mkdir(parents=True, exist_ok=True)

            # Sauvegarder le résumé complet
            results_file = matchings_dir / "results.json"
            with open(results_file, 'w', encoding='utf-8') as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)

            yield f"event: done\n"
            yield f"data: {json.dumps({'summary': summary})}\n\n"

        except Exception as e:
            import traceback
            yield f"event: error\n"
            yield f"data: {json.dumps({'code': 'MATCHING_ERROR', 'message': str(e), 'traceback': traceback.format_exc()})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@router.get("/results")
async def get_results(
    project_id: str = Query(..., description="ID du projet"),
    timestamp: str = Query(..., description="Timestamp du matching (format: YYYY-MM-DD_HH-MM-SS)")
):
    """
    Récupère les résultats d'un matching depuis le fichier results.json

    **Paramètres:**
    - `project_id`: ID du projet
    - `timestamp`: Timestamp du matching (ex: 2025-10-12_06-49-01)

    **Retourne:** Liste des CVs avec leurs scores et commentaires
    """
    from project_manager import ProjectManager

    pm = ProjectManager(projects_folder="projects")
    project_path = pm.get_project_path(project_id)

    if not project_path:
        raise HTTPException(
            status_code=404,
            detail=f"Projet {project_id} introuvable"
        )

    results_file = project_path / "matchings" / timestamp / "results.json"

    if not results_file.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Aucun résultat trouvé pour le projet {project_id} au timestamp {timestamp}"
        )

    try:
        with open(results_file, 'r', encoding='utf-8') as f:
            results_data = json.load(f)

        # Retourner les résultats directement
        return results_data.get('results', [])

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la lecture des résultats: {str(e)}"
        )


@router.get("/{project_id}/{timestamp}/results", response_model=MatchingResponse)
async def get_matching_results_by_params(project_id: str, timestamp: str):
    """
    Récupère les résultats d'un matching depuis l'historique

    **Path params:**
    - `project_id`: ID du projet
    - `timestamp`: Timestamp du matching
    """
    try:
        from unified_project_manager import UnifiedProjectManager

        pm = UnifiedProjectManager(enterprises_folder="enterprises")
        matching_data = pm.load_matching(project_id, timestamp)

        if not matching_data:
            raise HTTPException(
                status_code=404,
                detail=f"Matching {project_id}/{timestamp} introuvable"
            )

        # Nouveau format: results.json contient directement {"results": [...], "metadata": {...}}
        results = []
        results_data = matching_data.get("results", [])

        for cv_result in results_data:
            results.append(ResultatMatching(
                cv=cv_result.get("cv", "inconnu"),
                score_final=cv_result.get("score_final", 0.0),
                score_base=cv_result.get("score_base", 0.0),
                bonus_nice_have_multiplicateur=cv_result.get("bonus_nice_have_multiplicateur", 1.0),
                coefficient_qualite_experience=cv_result.get("coefficient_qualite_experience", 1.0),
                nice_have_manquants=cv_result.get("nice_have_manquants", []),
                commentaire_scoring=cv_result.get("commentaire_scoring", ""),
                appreciation_globale=cv_result.get("appreciation_globale", ""),
                evidences=cv_result.get("evidences"),
                evidence_map=cv_result.get("evidence_map"),
                flags=cv_result.get("flags"),
                original_cv_url=cv_result.get("original_cv_url")
            ))

        metadata_dict = matching_data.get("metadata", {})
        metadata = MatchingMetadata(
            total_cvs=metadata_dict.get("total_cvs", 0),
            filtered_must_have=metadata_dict.get("filtered_must_have", 0),
            top_reranked=metadata_dict.get("top_reranked", len(results)),
            duree_totale_s=metadata_dict.get("duree_totale_s", 0),
        )

        return MatchingResponse(results=results, metadata=metadata)

    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"Matching {project_id}/{timestamp} introuvable"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors du chargement du matching: {str(e)}"
        )


@router.get("/{matching_id}/results", response_model=MatchingResponse)
async def get_matching_results(matching_id: str):
    """
    Récupère les résultats d'un matching depuis l'historique

    **LEGACY ROUTE** - Format: {project_id}/{timestamp}

    **Note:** Utilisez plutôt `/matching/{project_id}/{timestamp}/results`
    """
    # Parser matching_id: format attendu "project_id/timestamp"
    parts = matching_id.split("/")
    if len(parts) != 2:
        raise HTTPException(
            status_code=400,
            detail="Format matching_id invalide. Attendu: {project_id}/{timestamp}"
        )

    project_id, timestamp = parts

    try:
        from unified_project_manager import UnifiedProjectManager

        pm = UnifiedProjectManager(enterprises_folder="enterprises")
        matching_data = pm.load_matching(project_id, timestamp)

        if not matching_data:
            raise HTTPException(
                status_code=404,
                detail=f"Matching {matching_id} introuvable"
            )

        # Nouveau format: results.json contient directement {"results": [...], "metadata": {...}}
        results = []
        results_data = matching_data.get("results", [])

        for cv_result in results_data:
            results.append(ResultatMatching(
                cv=cv_result.get("cv", "inconnu"),
                score_final=cv_result.get("score_final", 0.0),
                score_base=cv_result.get("score_base", 0.0),
                bonus_nice_have_multiplicateur=cv_result.get("bonus_nice_have_multiplicateur", 1.0),
                coefficient_qualite_experience=cv_result.get("coefficient_qualite_experience", 1.0),
                nice_have_manquants=cv_result.get("nice_have_manquants", []),
                commentaire_scoring=cv_result.get("commentaire_scoring", ""),
                appreciation_globale=cv_result.get("appreciation_globale", ""),
                evidences=cv_result.get("evidences"),
                evidence_map=cv_result.get("evidence_map"),
                flags=cv_result.get("flags"),
                original_cv_url=cv_result.get("original_cv_url")
            ))

        metadata_dict = matching_data.get("metadata", {})
        metadata = MatchingMetadata(
            total_cvs=metadata_dict.get("total_cvs", 0),
            filtered_must_have=metadata_dict.get("filtered_must_have", 0),
            top_reranked=metadata_dict.get("top_reranked", len(results)),
            duree_totale_s=metadata_dict.get("duree_totale_s", 0),
            timestamp=metadata_dict.get("timestamp", "")
        )

        return MatchingResponse(results=results, metadata=metadata)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la récupération du matching: {str(e)}"
        )


@router.get("/{project_id}/{timestamp}/export/csv")
async def export_matching_csv_new(project_id: str, timestamp: str):
    """
    Exporte les résultats en CSV

    **Path params:**
    - `project_id`: ID du projet
    - `timestamp`: Timestamp du matching (format: YYYY-MM-DD_HH-MM-SS)

    **Format CSV:**
    ```
    cv,score_final,score_base,bonus_nice_have,coefficient_experience,nice_have_manquants,commentaire_scoring
    cv1.pdf,0.85,0.80,0.95,1.15,"GraphQL;AWS",Excellent profil...
    ```
    """

    try:
        from unified_project_manager import UnifiedProjectManager
        import csv
        from io import StringIO

        pm = UnifiedProjectManager(enterprises_folder="enterprises")
        matching_data = pm.load_matching(project_id, timestamp)

        if not matching_data:
            raise HTTPException(
                status_code=404,
                detail=f"Matching {project_id}/{timestamp} introuvable"
            )

        # Générer CSV
        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow([
            "cv",
            "score_final",
            "score_base",
            "bonus_nice_have_multiplicateur",
            "coefficient_qualite_experience",
            "nice_have_manquants",
            "commentaire_scoring",
            "appreciation_globale",
            "evidences",
            "evidence_map",
            "flags"
        ])

        # Rows - nouveau format
        results = matching_data.get("results", [])
        for cv_result in results:
            nice_have_manquants = cv_result.get("nice_have_manquants", [])
            nice_have_str = ";".join(nice_have_manquants) if nice_have_manquants else ""

            # Formater evidences
            evidences_list = cv_result.get("evidences", [])
            evidences_str = ""
            if evidences_list:
                evidences_str = " | ".join([
                    f"{ev.get('id', '')}:{ev.get('type', '')}:{ev.get('ref', '')}"
                    for ev in evidences_list
                ])

            # Formater evidence_map (normaliser None → {})
            evidence_map = cv_result.get("evidence_map") or {}
            evidence_map_str = ""
            if evidence_map:
                parts = []
                if evidence_map.get("commentaire_scoring"):
                    parts.append(f"scoring:[{','.join(evidence_map['commentaire_scoring'])}]")
                if evidence_map.get("appreciation_globale"):
                    parts.append(f"globale:[{','.join(evidence_map['appreciation_globale'])}]")
                evidence_map_str = " | ".join(parts)

            # Formater flags
            flags = cv_result.get("flags", {})
            flags_str = ""
            if flags:
                parts = []
                gappes = flags.get("gappes", [])
                if gappes:
                    gaps_details = ", ".join([
                        f"{g.get('period', '')} ({g.get('duration_months', 0)}m)"
                        for g in gappes
                    ])
                    parts.append(f"Gaps: {gaps_details}")
                overlaps = flags.get("overlaps", [])
                if overlaps:
                    overlaps_details = ", ".join([
                        f"{o.get('overlap_period', '')} ({o.get('overlap_days', 0)}j)"
                        for o in overlaps
                    ])
                    parts.append(f"Overlaps: {overlaps_details}")
                flags_str = " | ".join(parts)

            writer.writerow([
                cv_result.get("cv", "inconnu"),
                round(cv_result.get("score_final", 0.0), 4),
                round(cv_result.get("score_base", 0.0), 4),
                round(cv_result.get("bonus_nice_have_multiplicateur", 1.0), 4),
                round(cv_result.get("coefficient_qualite_experience", 1.0), 4),
                nice_have_str,
                cv_result.get("commentaire_scoring", "").replace("\n", " "),
                cv_result.get("appreciation_globale", "").replace("\n", " "),
                evidences_str,
                evidence_map_str,
                flags_str
            ])

        # Retourner CSV
        from fastapi.responses import Response

        csv_content = output.getvalue()
        return Response(
            content=csv_content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=matching_{project_id}_{timestamp}.csv"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de l'export CSV: {str(e)}"
        )


@router.get("/{project_id}/{timestamp}/export/excel")
async def export_matching_excel_new(project_id: str, timestamp: str):
    """
    Exporte les résultats en Excel (XLSX) avec mise en forme

    **Path params:**
    - `project_id`: ID du projet
    - `timestamp`: Timestamp du matching (format: YYYY-MM-DD_HH-MM-SS)

    **Format Excel:**
    - Feuille "Résultats" avec colonnes :
      - Rang, CV, Score Final, Score Base, Malus Nice-have, Coeff. Qualité, Nice-have Manquants, Commentaire, Appréciation
    - Tri par score final décroissant
    - Mise en forme (couleurs, bordures, largeur colonnes)
    """

    try:
        from unified_project_manager import UnifiedProjectManager
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        pm = UnifiedProjectManager(enterprises_folder="enterprises")
        matching_data = pm.load_matching(project_id, timestamp)

        if not matching_data:
            raise HTTPException(
                status_code=404,
                detail=f"Matching {project_id}/{timestamp} introuvable"
            )

        # Créer workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats Matching"

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes
        headers = [
            "Rang",
            "CV",
            "Score Final",
            "Score Base",
            "Malus Nice-have (×)",
            "Coeff. Qualité (×)",
            "Nice-have Manquants",
            "Commentaire Scoring",
            "Appréciation Globale",
            "Evidences",
            "Evidence Map",
            "Flags"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # Récupérer et trier les résultats par score final décroissant
        results_data = matching_data.get("results", [])
        # Les résultats sont déjà triés par score final dans le fichier

        # Remplir les données
        for row_idx, cv_result in enumerate(results_data, start=2):
            nice_have_manquants = cv_result.get("nice_have_manquants", [])
            nice_have_str = ", ".join(nice_have_manquants) if nice_have_manquants else "Aucun"

            # Formater evidences
            evidences_list = cv_result.get("evidences", [])
            evidences_str = ""
            if evidences_list:
                evidences_str = "\n".join([
                    f"{ev.get('id', '')}: {ev.get('type', '')} - {ev.get('ref', '')}"
                    for ev in evidences_list
                ])

            # Formater evidence_map (normaliser None → {})
            evidence_map = cv_result.get("evidence_map") or {}
            evidence_map_str = ""
            if evidence_map:
                parts = []
                if evidence_map.get("commentaire_scoring"):
                    parts.append(f"Scoring: [{', '.join(evidence_map['commentaire_scoring'])}]")
                if evidence_map.get("appreciation_globale"):
                    parts.append(f"Globale: [{', '.join(evidence_map['appreciation_globale'])}]")
                evidence_map_str = "\n".join(parts)

            # Formater flags
            flags = cv_result.get("flags", {})
            flags_str = ""
            if flags:
                parts = []
                gappes = flags.get("gappes", [])
                if gappes:
                    parts.append(f"🔴 GAPS ({len(gappes)}):")
                    for g in gappes:
                        parts.append(f"  • {g.get('period', '')} ({g.get('duration_months', 0)} mois)")
                overlaps = flags.get("overlaps", [])
                if overlaps:
                    parts.append(f"⚠️ OVERLAPS ({len(overlaps)}):")
                    for o in overlaps:
                        same_company = " (même ent.)" if o.get('same_company') else ""
                        parts.append(f"  • {o.get('overlap_period', '')} ({o.get('overlap_days', 0)} jours){same_company}")
                flags_str = "\n".join(parts)

            row_data = [
                row_idx - 1,  # Rang
                cv_result.get("cv", "inconnu"),
                round(cv_result.get("score_final", 0.0), 4),
                round(cv_result.get("score_base", 0.0), 4),
                round(cv_result.get("bonus_nice_have_multiplicateur", 1.0), 4),
                round(cv_result.get("coefficient_qualite_experience", 1.0), 4),
                nice_have_str,
                cv_result.get("commentaire_scoring", ""),
                cv_result.get("appreciation_globale", ""),
                evidences_str,
                evidence_map_str,
                flags_str
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Coloration score final
                if col_idx == 3:  # Score Final
                    score_pct = value * 100
                    if score_pct >= 80:
                        cell.font = Font(bold=True, color="006100")
                    elif score_pct >= 60:
                        cell.font = Font(bold=True, color="0070C0")
                    elif score_pct >= 40:
                        cell.font = Font(color="FFC000")
                    else:
                        cell.font = Font(color="C00000")

        # Ajuster largeurs colonnes
        column_widths = [8, 40, 12, 12, 18, 18, 30, 50, 50, 40, 30, 40]
        for col_idx, width in enumerate(column_widths, start=1):
            col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + (col_idx // 26)) + chr(64 + (col_idx % 26))
            ws.column_dimensions[col_letter].width = width

        # Figer première ligne
        ws.freeze_panes = "A2"

        # Sauvegarder dans BytesIO
        excel_output = BytesIO()
        wb.save(excel_output)
        excel_output.seek(0)

        # Retourner fichier Excel
        from fastapi.responses import Response

        return Response(
            content=excel_output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=matching_{project_id}_{timestamp}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de l'export Excel: {str(e)}"
        )


@router.get("/{project_id}/{timestamp}/export/json", response_model=MatchingResponse)
async def export_matching_json_new(project_id: str, timestamp: str):
    """
    Exporte les résultats en JSON (même format que /results)

    **Path params:**
    - `project_id`: ID du projet
    - `timestamp`: Timestamp du matching (format: YYYY-MM-DD_HH-MM-SS)
    """
    # Réutiliser la fonction get_matching_results_by_params
    return await get_matching_results_by_params(project_id, timestamp)
